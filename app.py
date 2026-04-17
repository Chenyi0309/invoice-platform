import io
import os
import re
import time
from datetime import date, datetime, timedelta

import pandas as pd
import streamlit as st
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload, MediaIoBaseDownload
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# =========================
# Page config
# =========================
st.set_page_config(
    page_title="DSP Invoice Upload",
    page_icon="📄",
    layout="wide"
)

# =========================
# Custom CSS
# =========================
st.markdown("""
<style>
    .stApp {
        background: linear-gradient(180deg, #f8fbff 0%, #eef4ff 100%);
    }
    .hero-brand {
        font-size: 2.4rem;
        font-weight: 800;
        color: #111827;
        line-height: 1;
    }
    .hero-subtitle {
        font-size: 1rem;
        color: #475569;
        margin-top: 0.35rem;
        margin-bottom: 1.2rem;
    }
    .hero-badge {
        font-size: 1rem;
        font-weight: 700;
        color: #2563eb;
        background: #dbeafe;
        padding: 8px 14px;
        border-radius: 999px;
        display: inline-block;
        margin-bottom: 16px;
    }
    .main-card {
        background: white;
        padding: 28px 28px 24px 28px;
        border-radius: 22px;
        box-shadow: 0 10px 30px rgba(30, 41, 59, 0.08);
        border: 1px solid #e2e8f0;
        margin-bottom: 24px;
    }
    .section-title {
        font-size: 1.2rem;
        font-weight: 700;
        color: #0f172a;
        margin-bottom: 0.9rem;
    }
    .status-good {
        background: #ecfdf5;
        border: 1px solid #a7f3d0;
        color: #065f46;
        padding: 14px 16px;
        border-radius: 14px;
        font-weight: 700;
        margin-top: 12px;
        margin-bottom: 10px;
    }
    .status-bad {
        background: #fef2f2;
        border: 1px solid #fecaca;
        color: #991b1b;
        padding: 14px 16px;
        border-radius: 14px;
        font-weight: 700;
        margin-top: 12px;
        margin-bottom: 10px;
    }
    .status-warn {
        background: #fff7ed;
        border: 1px solid #fdba74;
        color: #9a3412;
        padding: 14px 16px;
        border-radius: 14px;
        font-weight: 700;
        margin-top: 12px;
        margin-bottom: 10px;
    }
    .metric-card {
        background: #ffffff;
        border: 1px solid #e5e7eb;
        border-radius: 16px;
        padding: 16px;
        box-shadow: 0 4px 14px rgba(15, 23, 42, 0.04);
    }
    .metric-title {
        color: #64748b;
        font-size: 0.92rem;
        margin-bottom: 6px;
    }
    .metric-value {
        color: #0f172a;
        font-size: 1.15rem;
        font-weight: 800;
        word-break: break-word;
    }
    .row-card {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 18px;
        padding: 16px;
        margin-bottom: 14px;
    }
    .stButton > button {
        width: 100%;
        background: linear-gradient(90deg, #2563eb 0%, #1d4ed8 100%);
        color: white;
        border: none;
        border-radius: 14px;
        padding: 0.82rem 1rem;
        font-weight: 700;
        font-size: 1rem;
        box-shadow: 0 8px 22px rgba(37, 99, 235, 0.22);
    }
    .stButton > button:hover {
        background: linear-gradient(90deg, #1d4ed8 0%, #1e40af 100%);
        color: white;
    }
    div[data-testid="stFileUploader"] {
        background: #f8fafc;
        border: 1.5px dashed #93c5fd;
        border-radius: 18px;
        padding: 8px 10px 2px 10px;
    }
    .footer-note {
        color: #64748b;
        font-size: 0.92rem;
        margin-top: 8px;
    }
</style>
""", unsafe_allow_html=True)

# =========================
# Config
# =========================
REGIONS = ["ORD", "IND", "CVG", "CMH", "MSP", "SDF", "LEX", "DTW", "CLE", "TOL", "STL", "OMA", "FWA"]

COLUMN_MAP = {
    "teamid": "team_id",
    "salary": "salary",
    "region": "warehouse",
    "dsp_name": "dsp_name",
}

SCOPES = ["https://www.googleapis.com/auth/drive"]

# =========================
# Secrets
# =========================
GOOGLE_CLIENT_ID = st.secrets["google_drive"]["client_id"]
GOOGLE_CLIENT_SECRET = st.secrets["google_drive"]["client_secret"]
GOOGLE_REFRESH_TOKEN = st.secrets["google_drive"]["refresh_token"]
GOOGLE_ROOT_FOLDER_ID = st.secrets["google_drive"]["root_folder_id"]

UPLOAD_ACCESS_CODE = st.secrets["app"]["upload_access_code"]
APP_TITLE = "UniUni • ORD Delivery Invoice"
APP_REGION_LABEL = st.secrets["app"].get("region_label", "Dispatch Upload")

# =========================
# Helpers
# =========================
def get_monday(d: date) -> date:
    return d - timedelta(days=d.weekday())


def monday_str(d: date) -> str:
    return get_monday(d).strftime("%Y%m%d")


def parse_yyyymmdd(s: str):
    try:
        return datetime.strptime(s, "%Y%m%d").date()
    except Exception:
        return None


def is_monday_string(s: str) -> bool:
    d = parse_yyyymmdd(s)
    return bool(d and d.weekday() == 0)


def clean_teamid(value) -> str:
    s = str(value).strip()
    s = re.sub(r"\.0$", "", s)
    return s


def normalize_money(v) -> float:
    if pd.isna(v):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    s = s.replace(",", "").replace("$", "")
    s = s.replace("(", "-").replace(")", "")
    s = re.sub(r"[^0-9.\-]", "", s)
    return float(s) if s not in ["", "-", "."] else 0.0


def get_extension(filename: str) -> str:
    _, ext = os.path.splitext(filename)
    return ext.lower() or ".pdf"


def sanitize_folder_name(name: str) -> str:
    name = str(name).strip().upper()
    return re.sub(r'[\\/:*?"<>|]', "_", name)


def format_currency(v):
    if v is None:
        return "-"
    return f"${v:,.2f}"


# =========================
# Google Drive Auth
# =========================
@st.cache_resource
def get_drive_service():
    creds = Credentials(
        token=None,
        refresh_token=GOOGLE_REFRESH_TOKEN,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        scopes=SCOPES,
    )
    creds.refresh(Request())
    return build("drive", "v3", credentials=creds, cache_discovery=False)


try:
    drive_service = get_drive_service()
except Exception as e:
    st.error(f"Google auth failed: {repr(e)}")
    st.stop()


# =========================
# Drive functions
# =========================
def get_root_folder():
    return {"id": GOOGLE_ROOT_FOLDER_ID, "name": "DSP_Invoices"}


def find_folder_by_name(name: str, parent_id: str = None):
    safe_name = name.replace("'", "\\'")
    if parent_id:
        query = (
            f"name = '{safe_name}' and mimeType = 'application/vnd.google-apps.folder' "
            f"and '{parent_id}' in parents and trashed = false"
        )
    else:
        query = (
            f"name = '{safe_name}' and mimeType = 'application/vnd.google-apps.folder' "
            f"and trashed = false"
        )

    results = drive_service.files().list(
        q=query,
        spaces="drive",
        fields="files(id, name)"
    ).execute()

    files = results.get("files", [])
    return files[0] if files else None


def create_folder(name: str, parent_id: str = None):
    metadata = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
    }
    if parent_id:
        metadata["parents"] = [parent_id]

    return drive_service.files().create(
        body=metadata,
        fields="id, name"
    ).execute()


def get_or_create_week_folder(week_monday: str):
    root = get_root_folder()
    folder = find_folder_by_name(week_monday, parent_id=root["id"])
    if folder:
        return folder
    return create_folder(week_monday, parent_id=root["id"])


def get_or_create_region_folder(week_monday: str, region: str):
    week_folder = get_or_create_week_folder(week_monday)
    safe_region = sanitize_folder_name(region)
    folder = find_folder_by_name(safe_region, parent_id=week_folder["id"])
    if folder:
        return folder
    return create_folder(safe_region, parent_id=week_folder["id"])


def find_file_in_folder(filename: str, folder_id: str):
    safe_filename = filename.replace("'", "\\'")
    query = f"name = '{safe_filename}' and '{folder_id}' in parents and trashed = false"
    results = drive_service.files().list(
        q=query,
        spaces="drive",
        fields="files(id, name, mimeType, createdTime)"
    ).execute()
    files = results.get("files", [])
    return files[0] if files else None


def list_uploaded_invoices(week_monday: str, region: str = None):
    if region:
        parent_folder = get_or_create_region_folder(week_monday, region)
    else:
        parent_folder = get_or_create_week_folder(week_monday)

    results = drive_service.files().list(
        q=f"'{parent_folder['id']}' in parents and trashed = false",
        spaces="drive",
        fields="files(id, name, mimeType, createdTime)"
    ).execute()

    files = results.get("files", [])
    invoice_files = []

    for f in files:
        name = f.get("name", "")
        if name == "Teams_merged.xlsx":
            continue
        if f.get("mimeType") == "application/vnd.google-apps.folder":
            continue
        invoice_files.append(f)

    invoice_files.sort(key=lambda x: x.get("name", ""))
    return invoice_files


def download_excel_from_drive(filename: str, folder_id: str) -> pd.DataFrame:
    file = find_file_in_folder(filename, folder_id)
    if not file:
        raise FileNotFoundError(f"{filename} not found in week folder.")

    last_error = None
    for attempt in range(3):
        try:
            request = drive_service.files().get_media(fileId=file["id"])
            buffer = io.BytesIO()
            downloader = MediaIoBaseDownload(buffer, request)

            done = False
            while not done:
                _, done = downloader.next_chunk()

            buffer.seek(0)
            return pd.read_excel(buffer)
        except Exception as e:
            last_error = e
            time.sleep(1.5 * (attempt + 1))

    raise RuntimeError(f"Failed to download {filename}: {last_error}")


def download_file_bytes_from_drive(filename: str, folder_id: str) -> bytes:
    file = find_file_in_folder(filename, folder_id)
    if not file:
        raise FileNotFoundError(f"{filename} not found in week folder.")

    request = drive_service.files().get_media(fileId=file["id"])
    buffer = io.BytesIO()
    downloader = MediaIoBaseDownload(buffer, request)

    done = False
    while not done:
        _, done = downloader.next_chunk()

    buffer.seek(0)
    return buffer.read()


def upload_file_to_drive(file_bytes: bytes, filename: str, folder_id: str):
    existing = find_file_in_folder(filename, folder_id)
    if existing:
        return "duplicate"

    file_metadata = {
        "name": filename,
        "parents": [folder_id],
    }

    ext = os.path.splitext(filename)[1].lower()
    mime_map = {
        ".pdf": "application/pdf",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".csv": "text/csv",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
    }
    mimetype = mime_map.get(ext, "application/octet-stream")

    media = MediaIoBaseUpload(
        io.BytesIO(file_bytes),
        mimetype=mimetype,
        resumable=False,
    )

    drive_service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id, name"
    ).execute()

    return "uploaded"


def mark_team_as_submitted(week_monday: str, teamid: str, region: str):
    week_folder = get_or_create_week_folder(week_monday)

    file_obj = find_file_in_folder("Teams_merged.xlsx", week_folder["id"])
    if not file_obj:
        raise FileNotFoundError("Teams_merged.xlsx not found in week folder.")

    excel_bytes = download_file_bytes_from_drive("Teams_merged.xlsx", week_folder["id"])
    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is not None:
            headers[str(val).strip()] = col

    team_col = headers.get(COLUMN_MAP["teamid"])
    region_col = headers.get(COLUMN_MAP["region"])

    if not team_col or not region_col:
        raise ValueError("Teams_merged.xlsx missing required columns for coloring.")

    fill = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")

    matched = False
    for row in range(2, ws.max_row + 1):
        excel_team = clean_teamid(ws.cell(row=row, column=team_col).value)
        excel_region = str(ws.cell(row=row, column=region_col).value).strip().upper()

        if excel_team == clean_teamid(teamid) and excel_region == str(region).strip().upper():
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill
            matched = True
            break

    if not matched:
        raise ValueError("Matching row not found in Teams_merged.xlsx.")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    drive_service.files().update(
        fileId=file_obj["id"],
        media_body=MediaIoBaseUpload(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=False
        )
    ).execute()


# =========================
# Business logic
# =========================
def load_weekly_teams(week_monday: str) -> pd.DataFrame:
    week_folder = get_or_create_week_folder(week_monday)
    df = download_excel_from_drive("Teams_merged.xlsx", week_folder["id"])
    df.columns = [str(c).strip() for c in df.columns]

    required = [COLUMN_MAP["teamid"], COLUMN_MAP["salary"], COLUMN_MAP["region"]]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Teams_merged.xlsx missing required columns: {missing}")

    df[COLUMN_MAP["teamid"]] = df[COLUMN_MAP["teamid"]].astype(str).map(clean_teamid)
    df[COLUMN_MAP["salary"]] = df[COLUMN_MAP["salary"]].map(normalize_money)
    df[COLUMN_MAP["region"]] = df[COLUMN_MAP["region"]].astype(str).str.strip().str.upper()

    return df


def get_expected_salary(df: pd.DataFrame, teamid: str, region: str):
    team_col = COLUMN_MAP["teamid"]
    salary_col = COLUMN_MAP["salary"]
    region_col = COLUMN_MAP["region"]

    subset = df[
        (df[team_col] == clean_teamid(teamid)) &
        (df[region_col] == str(region).strip().upper())
    ]

    if subset.empty:
        return None, None

    row = subset.iloc[0]
    return float(row[salary_col]), row


# =========================
# Session state for multi rows
# =========================
def init_invoice_rows():
    if "invoice_rows" not in st.session_state:
        st.session_state.invoice_rows = [
            {"teamid": "", "region": REGIONS[0], "file": None}
        ]


def add_invoice_row():
    st.session_state.invoice_rows.append(
        {"teamid": "", "region": REGIONS[0], "file": None}
    )


def remove_invoice_row(idx: int):
    if len(st.session_state.invoice_rows) > 1:
        st.session_state.invoice_rows.pop(idx)


# =========================
# Access control
# =========================
if "access_granted" not in st.session_state:
    st.session_state["access_granted"] = False

if not st.session_state["access_granted"]:
    st.markdown('<div class="main-card">', unsafe_allow_html=True)
    st.markdown(f'<div class="hero-brand">{APP_TITLE}</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="hero-subtitle">Enter upload access code / 请输入上传访问码</div>',
        unsafe_allow_html=True
    )
    code = st.text_input("Access Code / 访问码", type="password")
    if st.button("Enter / 进入"):
        if code == UPLOAD_ACCESS_CODE:
            st.session_state["access_granted"] = True
            st.rerun()
        else:
            st.error("Invalid code / 访问码错误")
    st.markdown('</div>', unsafe_allow_html=True)
    st.stop()


# =========================
# UI
# =========================
init_invoice_rows()
default_week = monday_str(date.today())

top1, top2 = st.columns([1, 6])
with top1:
    try:
        st.image("logo.png", width=110)
    except Exception:
        pass

with top2:
    st.markdown(f'<div class="hero-brand">{APP_TITLE}</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="hero-subtitle">Multi-site invoice upload: auto-match Teams_merged amount and save to the correct warehouse folder in Google Drive. / 多站点一次性上传：自动匹配 Teams_merged 金额，并保存到 Google Drive 对应仓库文件夹。</div>',
        unsafe_allow_html=True
    )

st.markdown('<div class="main-card">', unsafe_allow_html=True)
st.markdown('<div class="section-title">Batch Upload / 批量上传</div>', unsafe_allow_html=True)

input_week = st.text_input("Week Monday / 周一日期 (YYYYMMDD)", value=default_week)
st.markdown(
    f'<div class="hero-badge">One submission can include multiple warehouses / 一次提交可包含多个仓库</div>',
    unsafe_allow_html=True
)

teams_df = None
if is_monday_string(input_week):
    try:
        teams_df = load_weekly_teams(input_week)
    except Exception as e:
        st.markdown(
            f'<div class="status-warn">⚠️ Unable to load Teams_merged.xlsx for {input_week}: {e}</div>',
            unsafe_allow_html=True
        )
else:
    st.markdown(
        '<div class="status-warn">⚠️ Week Monday must be a Monday in YYYYMMDD format. / 日期必须是周一，格式为 YYYYMMDD。</div>',
        unsafe_allow_html=True
    )

row_summaries = []

for idx, row in enumerate(st.session_state.invoice_rows):
    st.markdown(f"### Site {idx + 1} / 站点 {idx + 1}")
    cols = st.columns([1.2, 1.2, 1.6, 0.8])

    with cols[0]:
        teamid = st.text_input(
            f"Team ID / 团队编号 #{idx + 1}",
            value=row.get("teamid", ""),
            key=f"teamid_{idx}",
            placeholder="例如 Example: 1206"
        )

    with cols[1]:
        current_region = row.get("region", REGIONS[0])
        region_index = REGIONS.index(current_region) if current_region in REGIONS else 0
        region = st.selectbox(
            f"Warehouse / 仓库 #{idx + 1}",
            REGIONS,
            index=region_index,
            key=f"region_{idx}"
        )

    with cols[2]:
        uploaded_file = st.file_uploader(
            f"Upload invoice file / 上传发票 #{idx + 1}",
            type=["pdf", "xlsx", "xls", "csv", "png", "jpg", "jpeg"],
            key=f"file_{idx}",
            help="支持拖拽上传 / Drag & drop supported",
        )

    with cols[3]:
        if st.button(f"Remove / 删除 #{idx + 1}", key=f"remove_{idx}"):
            remove_invoice_row(idx)
            st.rerun()

    expected_salary = None
    status = "Waiting"
    save_path = f"{input_week}/{region}" if input_week else f"-/{region}"

    if teams_df is not None and clean_teamid(teamid):
        expected_salary, _ = get_expected_salary(teams_df, clean_teamid(teamid), region)
        if expected_salary is not None:
            status = "Matched"
        else:
            status = "Not found in Teams_merged"
    elif not clean_teamid(teamid):
        status = "Team ID required"

    m1, m2, m3 = st.columns(3)
    with m1:
        st.markdown(
            f'<div class="metric-card"><div class="metric-title">Expected Amount / 应付金额</div><div class="metric-value">{format_currency(expected_salary)}</div></div>',
            unsafe_allow_html=True
        )
    with m2:
        st.markdown(
            f'<div class="metric-card"><div class="metric-title">Save Folder / 保存路径</div><div class="metric-value">{save_path}</div></div>',
            unsafe_allow_html=True
        )
    with m3:
        st.markdown(
            f'<div class="metric-card"><div class="metric-title">Status / 状态</div><div class="metric-value">{status}</div></div>',
            unsafe_allow_html=True
        )

    st.session_state.invoice_rows[idx] = {
        "teamid": teamid,
        "region": region,
        "file": uploaded_file,
    }

    row_summaries.append({
        "row_no": idx + 1,
        "teamid": clean_teamid(teamid),
        "region": region,
        "file": uploaded_file,
        "expected_salary": expected_salary,
        "status": status,
        "save_path": save_path,
    })

    st.markdown("---")

add_col1, add_col2 = st.columns([1, 3])
with add_col1:
    st.button("+ Add Another Site / 新增站点", on_click=add_invoice_row)

summary_df = pd.DataFrame([
    {
        "Row": item["row_no"],
        "Team ID": item["teamid"],
        "Warehouse": item["region"],
        "Expected Amount": format_currency(item["expected_salary"]),
        "File Selected": "Yes" if item["file"] is not None else "No",
        "Status": item["status"],
        "Save Folder": item["save_path"],
    }
    for item in row_summaries
])

st.subheader("Review / 提交预览")
st.dataframe(summary_df, use_container_width=True, hide_index=True)

validation_errors = []
seen_pairs = set()
for item in row_summaries:
    pair = (item["teamid"], item["region"])
    if not item["teamid"]:
        validation_errors.append(f"Row {item['row_no']}: Team ID is required.")
    if item["file"] is None:
        validation_errors.append(f"Row {item['row_no']}: Invoice file is required.")
    if teams_df is not None and item["expected_salary"] is None:
        validation_errors.append(f"Row {item['row_no']}: Team ID + Warehouse not found in Teams_merged.")
    if pair in seen_pairs:
        validation_errors.append(f"Row {item['row_no']}: Duplicate Team ID + Warehouse in same submission.")
    seen_pairs.add(pair)

if not is_monday_string(input_week):
    validation_errors.append("Week Monday must be a Monday in YYYYMMDD format.")

if teams_df is None:
    validation_errors.append("Teams_merged.xlsx is not available for this week.")

for err in validation_errors:
    st.markdown(f'<div class="status-warn">⚠️ {err}</div>', unsafe_allow_html=True)

submit_all = st.button(
    "Submit All Invoices / 一次性提交全部发票",
    type="primary",
    disabled=len(validation_errors) > 0
)

if submit_all and len(validation_errors) == 0:
    upload_results = []
    success_count = 0

    with st.spinner("Uploading invoices to Google Drive... / 正在上传发票到 Google Drive..."):
        for item in row_summaries:
            teamid = item["teamid"]
            region = item["region"]
            uploaded_file = item["file"]

            region_folder = get_or_create_region_folder(input_week, region)
            ext = get_extension(uploaded_file.name)
            new_filename = f"{teamid}{region}{input_week}{ext}"
            file_bytes = uploaded_file.read()

            try:
                result = upload_file_to_drive(file_bytes, new_filename, region_folder["id"])
                color_status = "Skipped"
                if result == "uploaded":
                    try:
                        mark_team_as_submitted(input_week, teamid, region)
                        color_status = "Colored"
                    except Exception as e:
                        color_status = f"Color failed: {e}"
                    success_count += 1

                upload_results.append({
                    "Row": item["row_no"],
                    "Team ID": teamid,
                    "Warehouse": region,
                    "Expected Amount": format_currency(item["expected_salary"]),
                    "Saved File": new_filename,
                    "Saved Folder": f"{input_week}/{region}",
                    "Upload Result": result,
                    "Teams_merged Update": color_status,
                })
            except Exception as e:
                upload_results.append({
                    "Row": item["row_no"],
                    "Team ID": teamid,
                    "Warehouse": region,
                    "Expected Amount": format_currency(item["expected_salary"]),
                    "Saved File": "-",
                    "Saved Folder": f"{input_week}/{region}",
                    "Upload Result": f"Failed: {e}",
                    "Teams_merged Update": "Not updated",
                })

    if success_count > 0:
        st.balloons()
        st.markdown(
            f'<div class="status-good">✅ {success_count} invoice(s) uploaded successfully. / 成功上传 {success_count} 个发票。</div>',
            unsafe_allow_html=True
        )

    result_df = pd.DataFrame(upload_results)
    st.subheader("Upload Result / 上传结果")
    st.dataframe(result_df, use_container_width=True, hide_index=True)

    # reset rows after full success or partial attempt
    st.session_state.invoice_rows = [{"teamid": "", "region": REGIONS[0], "file": None}]

st.markdown(
    '<div class="footer-note">Folder structure / 文件结构：DSP_Invoices / 周一日期 / 仓库 / 发票；Teams_merged.xlsx 保留在周文件夹根目录</div>',
    unsafe_allow_html=True
)

st.markdown("---")
st.subheader("Submitted Invoices / 已提交发票")

search_col1, search_col2 = st.columns([2, 1])
with search_col1:
    search_team = st.text_input(
        "Search by Team ID / 按 Team ID 搜索",
        value="",
        placeholder="例如 Example: 1363"
    )
with search_col2:
    search_region = st.selectbox("Search Warehouse / 搜索仓库", ["All"] + REGIONS)

try:
    if not search_team.strip():
        st.info("Please enter Team ID to search. / 请输入 Team ID 才会显示结果。")
    else:
        selected_region = None if search_region == "All" else search_region
        submitted_files = list_uploaded_invoices(input_week, selected_region)

        keyword = clean_teamid(search_team)
        submitted_files = [
            f for f in submitted_files
            if keyword in clean_teamid(f.get("name", ""))
        ]

        st.markdown(f"**Total matched / 匹配数量：{len(submitted_files)}**")

        if submitted_files:
            submitted_df = pd.DataFrame(
                {
                    "File Name / 文件名": [f["name"] for f in submitted_files],
                    "Created Time / 上传时间": [f.get("createdTime", "") for f in submitted_files],
                    "Warehouse Scope / 仓库范围": [selected_region if selected_region else "All under selected level" for _ in submitted_files],
                }
            )
            st.dataframe(submitted_df, use_container_width=True, hide_index=True)
        else:
            st.warning("No matching invoices found. / 没有找到匹配的发票。")

    if submitted_files:
        submitted_df = pd.DataFrame(
            {
                "File Name / 文件名": [f["name"] for f in submitted_files],
                "Created Time / 上传时间": [f.get("createdTime", "") for f in submitted_files],
                "Warehouse Scope / 仓库范围": [selected_region if selected_region else "All under selected level" for _ in submitted_files],
            }
        )
        st.dataframe(submitted_df, use_container_width=True, hide_index=True)
    else:
        st.info("No matching submitted invoices found for this week. / 本周没有符合条件的已提交发票。")

except Exception as e:
    st.warning(f"Failed to load submitted invoices: {e}")

st.markdown('</div>', unsafe_allow_html=True)
